import streamlit as st
import pandas as pd
import numpy as np
import joblib
import io
import warnings
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score

warnings.filterwarnings('ignore')

# ============================================================
#  CONFIGURACIÓN GENERAL
# ============================================================
st.set_page_config(
    page_title="StockSense | Predicción de Inventario",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

USERS = {
    "admin":    "admin123",
    "gerente":  "gerente2026",
    "analista": "analista2026",
}

# ============================================================
#  PALETA DE COLORES (referencia única, evita inconsistencias)
# ============================================================
C_BG          = "#0b1d33"
C_BG_SOFT     = "#11253f"
C_CARD        = "#16304f"
C_CARD_BORDER = "#22436b"
C_TEXT        = "#eef3fa"
C_TEXT_DIM    = "#9fb3cc"
C_ACCENT      = "#3b82f6"
C_ACCENT_SOFT = "#60a5fa"
C_DANGER      = "#f87171"
C_DANGER_BG   = "#3a1620"
C_WARNING     = "#fbbf24"
C_WARNING_BG  = "#3a2a10"
C_SUCCESS     = "#4ade80"
C_SUCCESS_BG  = "#0f3322"
C_WHITE_CARD  = "#ffffff"
C_DARK_TEXT   = "#0b1d33"

# ============================================================
#  ESTILOS CSS (exactamente iguales a tu versión)
# ============================================================
st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    html, body, [class*="css"] {{ font-family: 'Inter', sans-serif; }}

    .stApp {{
        background: linear-gradient(180deg, {C_BG} 0%, #0d2138 100%) !important;
    }}
    .block-container {{ padding-top: 1.2rem; padding-bottom: 3rem; max-width: 1280px; }}

    .stApp, .stApp p, .stApp span, .stApp label,
    .stApp .stMarkdown, .stApp .stText, .stApp li {{
        color: {C_TEXT} !important;
    }}
    h1, h2, h3, h4 {{ color: {C_TEXT} !important; font-weight: 800 !important; letter-spacing: -0.3px; }}
    .stCaption, [data-testid="stCaptionContainer"] {{ color: {C_TEXT_DIM} !important; }}

    .topbar {{
        display: flex; align-items: center; justify-content: space-between;
        padding: 14px 22px; background: {C_BG_SOFT};
        border-radius: 14px; margin-bottom: 18px;
        border: 1px solid {C_CARD_BORDER};
    }}
    .topbar-brand {{ display: flex; align-items: center; gap: 12px; }}
    .topbar-brand .logo-box {{
        width: 38px; height: 38px; border-radius: 10px;
        background: linear-gradient(135deg, {C_ACCENT} 0%, #1e40af 100%);
        display: flex; align-items: center; justify-content: center;
        font-weight: 800; color: white; font-size: 16px;
    }}
    .topbar-brand .brand-name {{ font-size: 17px; font-weight: 800; color: {C_TEXT}; }}
    .topbar-brand .brand-sub {{ font-size: 11px; color: {C_TEXT_DIM}; letter-spacing: 0.4px; }}
    .topbar-user {{
        font-size: 13px; color: {C_TEXT_DIM}; text-align: right;
    }}
    .topbar-user b {{ color: {C_TEXT}; }}

    .metric-card {{
        background: {C_WHITE_CARD};
        border-radius: 14px;
        padding: 20px 20px;
        text-align: center;
        margin-bottom: 12px;
        box-shadow: 0 6px 18px rgba(0,0,0,0.22);
        border-top: 4px solid #cbd5e1;
    }}
    .metric-card.danger  {{ border-top-color: #dc2626; }}
    .metric-card.warning {{ border-top-color: #d97706; }}
    .metric-card.success {{ border-top-color: #16a34a; }}
    .metric-card.info    {{ border-top-color: {C_ACCENT}; }}
    .metric-card .label {{
        font-size: 11.5px; color: #64748b !important; font-weight: 700;
        text-transform: uppercase; letter-spacing: 0.7px;
    }}
    .metric-card .value {{
        font-size: 29px; font-weight: 800; color: {C_DARK_TEXT} !important;
        margin: 7px 0 3px 0; line-height: 1.1;
    }}
    .metric-card .sub {{ font-size: 11.5px; color: #8a93a3 !important; font-weight: 500; }}

    .section-header {{
        background: linear-gradient(120deg, {C_ACCENT} 0%, #1e3a8a 100%);
        color: #ffffff !important; padding: 13px 20px; border-radius: 10px;
        font-size: 15px; font-weight: 700; margin: 24px 0 14px 0;
        letter-spacing: 0.2px;
    }}
    .section-header * {{ color: #ffffff !important; }}

    .subtle-header {{
        font-size: 14px; font-weight: 700; color: {C_TEXT_DIM} !important;
        text-transform: uppercase; letter-spacing: 0.6px;
        margin: 20px 0 10px 0; padding-bottom: 8px;
        border-bottom: 1px solid {C_CARD_BORDER};
    }}

    .alert-danger, .alert-warning, .alert-success {{
        padding: 13px 18px; border-radius: 10px; margin: 9px 0;
        font-weight: 600; font-size: 14px;
    }}
    .alert-danger  {{ background: {C_DANGER_BG};  border-left: 4px solid {C_DANGER};  color: #fecaca !important; }}
    .alert-warning {{ background: {C_WARNING_BG}; border-left: 4px solid {C_WARNING}; color: #fde68a !important; }}
    .alert-success {{ background: {C_SUCCESS_BG}; border-left: 4px solid {C_SUCCESS}; color: #bbf7d0 !important; }}
    .alert-danger *, .alert-warning *, .alert-success * {{ color: inherit !important; }}

    .product-row {{
        background: {C_CARD};
        border-radius: 10px;
        padding: 13px 18px;
        margin-bottom: 7px;
        border: 1px solid {C_CARD_BORDER};
        border-left: 4px solid #475569;
    }}
    .product-row.danger  {{ border-left-color: {C_DANGER}; }}
    .product-row.warning {{ border-left-color: {C_WARNING}; }}
    .product-row.success {{ border-left-color: {C_SUCCESS}; }}
    .product-row b {{ color: {C_TEXT} !important; }}
    .product-row span {{ color: {C_TEXT_DIM} !important; }}

    .repo-card {{
        background: {C_CARD};
        border: 1px solid {C_CARD_BORDER};
        border-radius: 12px;
        padding: 16px 18px;
        margin-bottom: 10px;
        display: flex; align-items: center; justify-content: space-between;
    }}
    .repo-date-box {{
        min-width: 76px; text-align: center; padding: 8px 10px;
        border-radius: 10px; font-weight: 800; margin-right: 16px;
    }}
    .repo-date-box .day {{ font-size: 20px; line-height: 1; }}
    .repo-date-box .mon {{ font-size: 10px; letter-spacing: 0.5px; opacity: 0.85; }}

    .demo-badge {{
        display: inline-block; background: {C_WARNING_BG}; color: {C_WARNING} !important;
        font-size: 11px; font-weight: 700; padding: 4px 10px; border-radius: 20px;
        border: 1px solid {C_WARNING}; letter-spacing: 0.3px; margin-bottom: 10px;
    }}

    .login-card-icon {{ text-align: center; margin-bottom: 16px; }}
    .login-box {{
        background: {C_BG_SOFT}; border: 1px solid {C_CARD_BORDER};
        border-radius: 18px; padding: 36px 34px;
    }}
    .login-title {{
        text-align: center; font-size: 23px; font-weight: 800;
        color: {C_TEXT} !important; margin-bottom: 4px;
    }}
    .login-sub {{
        text-align: center; font-size: 13px; color: {C_TEXT_DIM} !important;
        margin-bottom: 26px;
    }}
    .login-footer-box {{
        text-align: center; margin-top: 18px; padding: 12px;
        background: rgba(255,255,255,0.04); border-radius: 10px;
        font-size: 11.5px; color: {C_TEXT_DIM} !important; line-height: 1.8;
    }}

    .stTextInput input, .stNumberInput input {{
        background-color: {C_BG_SOFT} !important;
        color: {C_TEXT} !important;
        border: 1px solid {C_CARD_BORDER} !important;
        border-radius: 9px !important;
    }}
    .stSelectbox > div > div {{
        background-color: {C_BG_SOFT} !important;
        border: 1px solid {C_CARD_BORDER} !important;
        border-radius: 9px !important;
        color: {C_TEXT} !important;
    }}
    .stSelectbox div[data-baseweb="select"] span {{ color: {C_TEXT} !important; }}
    div[data-baseweb="popover"] li {{
        background-color: {C_BG_SOFT} !important;
        color: {C_TEXT} !important;
    }}
    div[data-baseweb="popover"] li:hover {{
        background-color: {C_CARD} !important;
    }}

    .stButton > button {{
        border-radius: 9px !important; font-weight: 700 !important;
        padding: 9px 18px !important; border: none !important;
    }}
    .stButton > button[kind="primary"] {{
        background: linear-gradient(120deg, {C_ACCENT} 0%, #1e3a8a 100%) !important;
        color: #ffffff !important;
        box-shadow: 0 4px 14px rgba(59,130,246,0.35) !important;
    }}
    .stButton > button[kind="secondary"] {{
        background: {C_CARD} !important;
        color: {C_TEXT} !important;
        border: 1px solid {C_CARD_BORDER} !important;
    }}

    .stTabs [data-baseweb="tab-list"] {{
        gap: 4px; background: {C_BG_SOFT}; padding: 6px;
        border-radius: 12px; border: 1px solid {C_CARD_BORDER};
    }}
    .stTabs [data-baseweb="tab"] {{
        border-radius: 8px; font-weight: 700; color: {C_TEXT_DIM} !important;
        padding: 10px 22px; font-size: 14px;
    }}
    .stTabs [data-baseweb="tab"] p {{ color: inherit !important; }}
    .stTabs [aria-selected="true"] {{
        background: {C_ACCENT} !important; color: #ffffff !important;
    }}
    .stTabs [aria-selected="true"] p {{ color: #ffffff !important; }}

    [data-testid="stDataFrame"] {{
        border-radius: 12px; overflow: hidden; border: 1px solid {C_CARD_BORDER};
    }}

    section[data-testid="stSidebar"] {{
        background: #081628 !important;
        border-right: 1px solid {C_CARD_BORDER};
    }}
    section[data-testid="stSidebar"] * {{ color: {C_TEXT} !important; }}
    section[data-testid="stSidebar"] hr {{ border-color: {C_CARD_BORDER}; }}
    section[data-testid="stSidebar"] .stButton > button {{
        background: {C_CARD} !important;
        color: {C_TEXT} !important;
        border: 1px solid {C_CARD_BORDER} !important;
        width: 100%;
    }}
    section[data-testid="stSidebar"] .stButton > button:hover {{
        background: {C_DANGER} !important;
        border-color: {C_DANGER} !important;
        color: #ffffff !important;
    }}

    [data-testid="stAlert"] {{ border-radius: 10px; }}

    .footer-credits {{
        text-align: center; font-size: 11px; color: {C_TEXT_DIM} !important;
        margin-top: 46px; padding-top: 16px;
        border-top: 1px solid {C_CARD_BORDER};
    }}

    footer {{ visibility: hidden; }}
    #MainMenu {{ visibility: hidden; }}
    header[data-testid="stHeader"] {{ background: transparent; }}
</style>
""", unsafe_allow_html=True)


# ============================================================
#  HELPERS DE UI
# ============================================================

def metric_card(label, value, sub="", tone="info"):
    st.markdown(f"""
    <div class="metric-card {tone}">
        <div class="label">{label}</div>
        <div class="value">{value}</div>
        <div class="sub">{sub}</div>
    </div>
    """, unsafe_allow_html=True)


def section_header(title):
    st.markdown(f'<div class="section-header">{title}</div>', unsafe_allow_html=True)


def subtle_header(title):
    st.markdown(f'<div class="subtle-header">{title}</div>', unsafe_allow_html=True)


def footer_credits():
    st.markdown("""
    <div class="footer-credits">
        StockSense &nbsp;·&nbsp; Plataforma de predicción de demanda e inventario<br>
        Desarrollado por Sara Camila Mayorca Parra &amp; Luis Alejandro Palomino Acuña
    </div>
    """, unsafe_allow_html=True)


def configurar_matplotlib_oscuro():
    plt.rcParams.update({
        'figure.facecolor':  C_CARD,
        'axes.facecolor':    C_CARD,
        'savefig.facecolor': C_CARD,
        'axes.edgecolor':    C_CARD_BORDER,
        'axes.labelcolor':   C_TEXT,
        'xtick.color':       C_TEXT_DIM,
        'ytick.color':       C_TEXT_DIM,
        'text.color':        C_TEXT,
        'axes.titlecolor':   C_TEXT,
        'grid.color':        '#22436b',
        'font.family':       'sans-serif',
    })

configurar_matplotlib_oscuro()


# ============================================================
#  MOTOR DE DATOS Y MODELO (PRE-ENTRENADO)
# ============================================================

@st.cache_resource(show_spinner=False)
def cargar_modelo_y_datos():
    """Carga el modelo entrenado, scaler, encoder y dataset histórico.
    Estos archivos deben estar en el mismo directorio que la app.
    Se obtienen desde el notebook de la tesis con:
        joblib.dump(modelo, 'modelo_rf_optimizado.pkl')
        joblib.dump(scaler, 'scaler.pkl')
        joblib.dump(le_country, 'le_country.pkl')
        joblib.dump(features, 'features.pkl')
        df_agg.to_csv('df_agg_online_retail.csv', index=False)
    """
    modelo  = joblib.load('modelo_rf_optimizado.pkl')
    scaler  = joblib.load('scaler.pkl')
    le_pais = joblib.load('le_country.pkl')
    features = joblib.load('features.pkl')  # lista exacta: ['UnitPrice','Country_Code','Month',...]

    # El dataset agregado ya contiene las columnas necesarias
    df_agg = pd.read_csv('df_agg_online_retail.csv')
    # Renombrar columnas a nombres en español para la interfaz
    df_agg = df_agg.rename(columns={
        'UnitPrice': 'Precio',
        'Country':   'Pais',        # asumimos que existe la columna Country
        'Quantity':  'Cantidad',
        'Transacciones': 'Pedidos',
        'Month':     'Mes',
        'Quarter':   'Trimestre',
        'DayOfWeek': 'DiaSemana',
        'Hour':      'Hora',
        'Year':      'Anio',
        'Week':      'Semana'
    })
    # Si el dataset no tiene 'Pais', lo creamos a partir del código
    if 'Pais' not in df_agg.columns and 'Country_Code' in df_agg.columns:
        # Intentar mapear de vuelta con le_pais
        df_agg['Pais'] = le_pais.inverse_transform(df_agg['Country_Code'])

    # Métricas finales del modelo (obtenidas del notebook)
    metricas = {
        'mae': 24.28,
        'rmse': 92.58,
        'r2': 0.29,
        'confiabilidad': 73,  # %
    }

    return {
        'modelo': modelo,
        'scaler': scaler,
        'le_pais': le_pais,
        'features': features,
        'df_agg': df_agg,
        'metricas': metricas,
    }


def predecir_producto(payload, producto, pais=None, semanas_adelante=1):
    """Predice la cantidad total de unidades a vender en las próximas N semanas
    para un producto y país dados. Utiliza el modelo pre‑entrenado exactamente
    con las 9 variables que se usaron en la tesis."""
    df_agg  = payload['df_agg']
    modelo  = payload['modelo']
    scaler  = payload['scaler']
    le_pais = payload['le_pais']
    features = payload['features']  # ['UnitPrice','Country_Code','Month','Quarter','Week','DayOfWeek','Hour','Year','Transacciones']

    hist = df_agg[df_agg['Producto'] == producto]
    if len(hist) == 0:
        return None

    # Si no se especifica país, usar el más frecuente en el historial del producto
    if pais is None:
        pais = hist['Pais'].mode().iloc[0]

    # Calcular promedios del producto
    precio_prom   = hist['Precio'].mean()
    pedidos_prom  = hist['Pedidos'].mean()          # antes llamado Transacciones
    hora_prom     = hist['Hora'].mean() if 'Hora' in hist else 12  # hora media de pedidos

    # Última semana registrada
    ultima_semana = hist['Semana'].max()
    anio_actual   = hist['Anio'].max()

    # Codificar país
    try:
        pais_code = le_pais.transform([pais])[0]
    except:
        pais_code = 0

    # Acumular predicción semana a semana (igual que en el notebook)
    total_predicho = 0
    for paso in range(1, semanas_adelante + 1):
        semana_obj = ultima_semana + paso
        anio_obj   = anio_actual
        if semana_obj > 52:
            semana_obj -= 52
            anio_obj   += 1
        # Estimar mes y trimestre a partir de la semana (aproximado)
        mes_obj  = min(12, max(1, round(semana_obj / 4.33)))
        trim_obj = (mes_obj - 1) // 3 + 1
        # Asumimos jueves (3) como día típico de compra
        dia_sem  = 3

        entrada = pd.DataFrame([[
            precio_prom, pais_code, mes_obj, trim_obj,
            semana_obj, dia_sem, hora_prom, anio_obj, pedidos_prom
        ]], columns=features)
        entrada_s = scaler.transform(entrada)
        total_predicho += max(0, modelo.predict(entrada_s)[0])

    return round(total_predicho)


def calcular_estado_inventario(payload, producto, stock_actual, pais=None, semanas_adelante=1):
    """Calcula el estado del inventario (crítico, atención, exceso, óptimo)
    para un producto, usando la predicción semanal y el stock de seguridad."""
    df_agg = payload['df_agg']
    hist = df_agg[df_agg['Producto'] == producto]
    demanda_pred = predecir_producto(payload, producto, pais, semanas_adelante)
    if demanda_pred is None:
        return None

    # Desviación estándar histórica para el stock de seguridad
    std_hist = hist['Cantidad'].std()
    if pd.isna(std_hist) or std_hist == 0:
        std_hist = hist['Cantidad'].mean() * 0.3

    # Stock de seguridad (factor 1.65 para nivel de servicio del 95%)
    stock_seguridad  = round(1.65 * std_hist * (semanas_adelante ** 0.5))
    punto_reposicion = demanda_pred + stock_seguridad
    demanda_diaria   = demanda_pred / (7 * semanas_adelante) if demanda_pred > 0 else 0.01
    dias_cobertura   = round(stock_actual / demanda_diaria) if demanda_diaria > 0 else 999

    if stock_actual < stock_seguridad:
        estado, unidades_sugeridas = 'critico', max(0, punto_reposicion + demanda_pred - stock_actual)
    elif stock_actual < punto_reposicion:
        estado, unidades_sugeridas = 'atencion', max(0, punto_reposicion - stock_actual)
    elif stock_actual > punto_reposicion * 1.8:
        estado, unidades_sugeridas = 'exceso', 0
    else:
        estado, unidades_sugeridas = 'optimo', 0

    fecha_sugerida = datetime.now() + timedelta(days=max(1, dias_cobertura - 2))
    return {
        'demanda_predicha': demanda_pred,
        'stock_seguridad': stock_seguridad,
        'punto_reposicion': round(punto_reposicion),
        'dias_cobertura': dias_cobertura,
        'estado': estado,
        'unidades_sugeridas': round(unidades_sugeridas),
        'fecha_sugerida_dt': fecha_sugerida,
        'fecha_sugerida': fecha_sugerida.strftime('%d/%m/%Y'),
        'semanas_adelante': semanas_adelante,
    }


# ============================================================
#  REPORTES (Excel, CSV, TXT) – se mantienen igual
# ============================================================
def generar_reporte_excel(df_reporte):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_reporte.to_excel(writer, index=False, sheet_name='Inventario', startrow=1)
        ws = writer.sheets['Inventario']
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_reporte.columns))
        titulo_cell = ws.cell(row=1, column=1)
        titulo_cell.value = f"StockSense · Reporte de inventario — {datetime.now().strftime('%d/%m/%Y')}"
        titulo_cell.font = Font(bold=True, color='FFFFFF', size=12)
        titulo_cell.fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24
        header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col_idx in range(1, len(df_reporte.columns) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        estado_colors = {'Crítico': 'FCA5A5', 'Atención': 'FDE68A',
                         'Exceso': 'BFDBFE', 'Óptimo': 'BBF7D0'}
        if 'Estado' in df_reporte.columns:
            estado_col_idx = df_reporte.columns.get_loc('Estado') + 1
            for row_idx in range(3, len(df_reporte) + 3):
                estado_val = ws.cell(row=row_idx, column=estado_col_idx).value
                color = estado_colors.get(estado_val)
                if color:
                    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    for col_idx in range(1, len(df_reporte.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill
        for col_idx, col_name in enumerate(df_reporte.columns, start=1):
            max_len = max(df_reporte[col_name].astype(str).map(len).max(), len(str(col_name))) + 3
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 45)
        ws.freeze_panes = 'A3'
    buffer.seek(0)
    return buffer


def generar_resumen_texto(df_reporte):
    criticos = df_reporte[df_reporte['Estado'] == 'Crítico']
    exceso   = df_reporte[df_reporte['Estado'] == 'Exceso']
    total_demanda = df_reporte['Demanda estimada (próx. semana)'].sum()
    lineas_criticos = chr(10).join(
        '- ' + str(r['Producto'])[:60] + f" (pedir {r['Unidades sugeridas a pedir']} unidades)"
        for _, r in criticos.head(15).iterrows())
    lineas_exceso = chr(10).join('- ' + str(r['Producto'])[:60] for _, r in exceso.head(15).iterrows())
    return f"""RESUMEN EJECUTIVO DE INVENTARIO
Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}
{'='*50}

VENTAS ESTIMADAS PROXIMA SEMANA: {total_demanda:,.0f} unidades

PRODUCTOS EN RIESGO CRITICO: {len(criticos)}
{lineas_criticos}

PRODUCTOS CON SOBRESTOCK: {len(exceso)}
{lineas_exceso}

{'='*50}
Reporte generado por StockSense
"""


# ============================================================
#  LOGIN
# ============================================================
def show_login():
    col1, col2, col3 = st.columns([1, 1.3, 1])
    with col2:
        st.markdown('<br><br>', unsafe_allow_html=True)
        st.markdown('<div class="login-box">', unsafe_allow_html=True)
        st.markdown("""
        <div class="login-card-icon"><span style="font-size:48px;">📦</span></div>
        <div class="login-title">StockSense</div>
        <div class="login-sub">Predicción inteligente de demanda e inventario</div>
        """, unsafe_allow_html=True)

        with st.form("login_form"):
            usuario  = st.text_input("Usuario", placeholder="Ingresa tu usuario")
            password = st.text_input("Contraseña", type="password", placeholder="Ingresa tu contraseña")
            submit   = st.form_submit_button("Ingresar", use_container_width=True, type="primary")

        if submit:
            if usuario in USERS and USERS[usuario] == password:
                st.session_state['logged_in'] = True
                st.session_state['usuario'] = usuario
                st.rerun()
            else:
                st.error("Usuario o contraseña incorrectos.")

        st.markdown("""
        <div class="login-footer-box">
            <b>Acceso de demostración</b><br>
            admin / admin123 &nbsp;·&nbsp; gerente / gerente2026 &nbsp;·&nbsp; analista / analista2026
        </div>
        """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)


# ============================================================
#  TOPBAR Y SIDEBAR
# ============================================================
def show_topbar():
    usuario = st.session_state.get('usuario', '')
    st.markdown(f"""
    <div class="topbar">
        <div class="topbar-brand">
            <div class="logo-box">S</div>
            <div>
                <div class="brand-name">StockSense</div>
                <div class="brand-sub">INVENTARIO INTELIGENTE</div>
            </div>
        </div>
        <div class="topbar-user">Sesión activa<br><b>{usuario.upper()}</b></div>
    </div>
    """, unsafe_allow_html=True)


def show_sidebar():
    with st.sidebar:
        st.markdown("""
        <div style="text-align:center; padding:10px 0 18px;">
            <div style="font-size:34px;">📦</div>
            <div style="font-size:15px; font-weight:800; margin-top:6px;">StockSense</div>
        </div>
        <hr>
        """, unsafe_allow_html=True)
        st.markdown("**Cuenta**")
        st.caption(st.session_state.get('usuario', '').upper())

        st.markdown("<br>", unsafe_allow_html=True)
        # Botón para recargar el modelo (por si hay actualizaciones)
        if st.button("Recargar modelo", use_container_width=True):
            st.cache_resource.clear()
            st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("<hr>", unsafe_allow_html=True)
        if st.button("Cerrar sesión", use_container_width=True):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()


# ============================================================
#  VISTA: INICIO
# ============================================================
def view_inicio():
    payload = st.session_state['payload']
    df_agg  = payload['df_agg']
    productos = df_agg['Producto'].unique()

    if 'stock_data' not in st.session_state:
        st.session_state['stock_data'] = {}

    productos_muestra = productos[:60] if len(productos) > 60 else productos
    estados = []
    for p in productos_muestra:
        stock_actual = st.session_state['stock_data'].get(p)
        if stock_actual is None:
            hist_prod = df_agg[df_agg['Producto'] == p]['Cantidad']
            stock_actual = int(hist_prod.mean() * 2) if len(hist_prod) else 50
            st.session_state['stock_data'][p] = stock_actual
        info = calcular_estado_inventario(payload, p, stock_actual)
        if info:
            info['producto'], info['stock_actual'] = p, stock_actual
            estados.append(info)

    criticos = [e for e in estados if e['estado'] == 'critico']
    atencion = [e for e in estados if e['estado'] == 'atencion']
    exceso   = [e for e in estados if e['estado'] == 'exceso']
    demanda_total = sum(e['demanda_predicha'] for e in estados)

    st.caption(f"Modelo entrenado con Online Retail 2010-2011 | "
              f"Confiabilidad: {payload['metricas']['confiabilidad']:.0f}%")

    with st.expander("📘 ¿Cómo leer este panel?"):
        st.markdown("""
        - **Por agotarse**: el stock está por debajo del mínimo de seguridad; hay que reponer ya.
        - **Necesitan atención**: llegarán a su punto de reposición pronto; conviene planificar el pedido.
        - **Con exceso**: tienes más stock del que normalmente necesitas; evalúa una promoción o pausar compras.
        - **Ventas próx. semana**: unidades que el modelo estima que venderás en los próximos 7 días, en total.
        """)

    col1, col2, col3, col4 = st.columns(4)
    with col1: metric_card("Por agotarse", f"{len(criticos)}", "riesgo crítico", "danger")
    with col2: metric_card("Necesitan atención", f"{len(atencion)}", "bajo nivel ideal", "warning")
    with col3: metric_card("Con exceso", f"{len(exceso)}", "sobrestock", "info")
    with col4: metric_card("Ventas próx. semana", f"{demanda_total:,.0f}", "unidades estimadas", "success")

    section_header("¿Qué productos se te van a acabar?")
    if criticos:
        for e in sorted(criticos, key=lambda x: x['dias_cobertura'])[:8]:
            col_a, col_b = st.columns([4, 1])
            with col_a:
                st.markdown(f"""
                <div class="product-row danger">
                    <b>{e['producto'][:60]}</b><br>
                    <span style="font-size:13px;">
                        Se agota en {e['dias_cobertura']} días · Stock actual: {e['stock_actual']} ·
                        Sugerido pedir <b>{e['unidades_sugeridas']} unidades</b> antes del {e['fecha_sugerida']}
                    </span>
                </div>""", unsafe_allow_html=True)
            with col_b:
                if st.button("Ver detalle", key=f"crit_{e['producto']}", use_container_width=True):
                    st.session_state['producto_detalle'] = e['producto']
                    st.session_state['_force_inventario'] = True
                    st.rerun()
    else:
        st.markdown('<div class="alert-success">No tienes productos en riesgo crítico de desabastecimiento.</div>',
                   unsafe_allow_html=True)

    section_header("¿Qué productos te están sobrando?")
    if exceso:
        for e in exceso[:6]:
            st.markdown(f"""
            <div class="product-row warning">
                <b>{e['producto'][:60]}</b><br>
                <span style="font-size:13px;">
                    Stock actual: {e['stock_actual']} unidades · Demanda estimada: {e['demanda_predicha']} unidades/semana
                </span>
            </div>""", unsafe_allow_html=True)
    else:
        st.markdown('<div class="alert-success">No se detecta sobrestock relevante en tu inventario.</div>',
                   unsafe_allow_html=True)

    section_header("¿Cuánto vas a vender la próxima semana?")
    top_demanda = sorted(estados, key=lambda x: -x['demanda_predicha'])[:6]
    df_top = pd.DataFrame([{'Producto': e['producto'][:38], 'Unidades': e['demanda_predicha']}
                           for e in top_demanda])
    fig, ax = plt.subplots(figsize=(10, 3.6))
    ax.barh(df_top['Producto'][::-1], df_top['Unidades'][::-1],
            color=C_ACCENT_SOFT, edgecolor=C_BG, linewidth=0.5, height=0.6)
    ax.set_title('Productos con mayor demanda estimada', fontweight='bold', fontsize=12, pad=10)
    ax.spines[['top', 'right']].set_visible(False)
    for i, v in enumerate(df_top['Unidades'][::-1]):
        ax.text(v + max(df_top['Unidades'])*0.02, i, f'{v:.0f}', va='center',
               fontsize=9, color=C_TEXT, fontweight='bold')
    plt.tight_layout()
    st.pyplot(fig)
    plt.close()

    footer_credits()


# ============================================================
#  VISTA: MI INVENTARIO
# ============================================================
def view_inventario():
    payload = st.session_state['payload']
    df_agg  = payload['df_agg']
    productos = sorted(df_agg['Producto'].unique())

    if 'stock_data' not in st.session_state:
        st.session_state['stock_data'] = {}

    preseleccionado = st.session_state.pop('producto_detalle', None)

    subtle_header("Buscar un producto")
    col1, col2, col3 = st.columns([2.4, 1, 1])
    with col1:
        idx = productos.index(preseleccionado) if preseleccionado in productos else 0
        producto = st.selectbox("Producto", productos, index=idx, label_visibility="collapsed")
    with col2:
        valor_default = st.session_state['stock_data'].get(
            producto, int(df_agg[df_agg['Producto']==producto]['Cantidad'].mean()*2))
        stock_actual = st.number_input("Stock actual", min_value=0, max_value=1000000,
                                       value=valor_default, step=1)
        st.session_state['stock_data'][producto] = stock_actual
    with col3:
        horizonte = st.selectbox("Horizonte", [1, 2, 3, 4],
                                 format_func=lambda x: f"{x} semana(s)")

    # Obtener país por defecto del producto
    pais_default = df_agg[df_agg['Producto']==producto]['Pais'].mode().iloc[0]
    pais = st.selectbox("País", sorted(df_agg['Pais'].unique()),
                        index=list(sorted(df_agg['Pais'].unique())).index(pais_default))

    info = calcular_estado_inventario(payload, producto, stock_actual, pais, semanas_adelante=horizonte)
    if info is None:
        st.warning("No hay suficiente historial para este producto.")
        return

    estado_labels = {
        'critico':  ('Riesgo crítico de desabastecimiento', 'alert-danger'),
        'atencion': ('Requiere reposición pronto', 'alert-warning'),
        'exceso':   ('Sobrestock detectado', 'alert-warning'),
        'optimo':   ('Nivel óptimo', 'alert-success'),
    }
    texto, clase = estado_labels[info['estado']]
    st.markdown(f'<div class="{clase}">{texto}</div>', unsafe_allow_html=True)

    col1, col2, col3, col4 = st.columns(4)
    with col1: metric_card("Demanda estimada", f"{info['demanda_predicha']:,}", f"unidades en {horizonte} semana(s)")
    with col2: metric_card("Stock de seguridad", f"{info['stock_seguridad']:,}", "mínimo recomendado")
    with col3: metric_card("Punto de reposición", f"{info['punto_reposicion']:,}", "umbral para pedir")
    with col4: metric_card("Cobertura actual", f"{info['dias_cobertura']} días", "con el stock de hoy")

    if info['unidades_sugeridas'] > 0:
        st.markdown(f"""
        <div class="alert-warning">
            Te recomendamos pedir <b>{info['unidades_sugeridas']} unidades</b>
            antes del <b>{info['fecha_sugerida']}</b> para evitar quiebre de stock.
        </div>""", unsafe_allow_html=True)

    section_header(f"Histórico de ventas — {producto[:50]}")
    hist = df_agg[df_agg['Producto'] == producto].sort_values(['Anio', 'Semana'])
    if len(hist) > 1:
        fig, ax = plt.subplots(figsize=(12, 4))
        x_labels = [f"S{int(s)}-{int(a)}" for s, a in zip(hist['Semana'], hist['Anio'])]
        ax.plot(range(len(hist)), hist['Cantidad'], color=C_ACCENT_SOFT,
               linewidth=2.2, marker='o', markersize=4, markerfacecolor=C_ACCENT_SOFT)
        ax.fill_between(range(len(hist)), hist['Cantidad'], alpha=0.15, color=C_ACCENT_SOFT)
        ax.axhline(info['demanda_predicha'] / horizonte, color=C_SUCCESS, linestyle='--', linewidth=2,
                  label=f"Promedio semanal estimado: {info['demanda_predicha'] / horizonte:.0f}")
        step = max(1, len(x_labels)//12)
        ax.set_xticks(range(0, len(x_labels), step))
        ax.set_xticklabels(x_labels[::step], rotation=40, fontsize=8)
        ax.set_ylabel('Unidades vendidas')
        ax.spines[['top', 'right']].set_visible(False)
        leg = ax.legend(facecolor=C_CARD, edgecolor=C_CARD_BORDER, fontsize=9)
        for text in leg.get_texts(): text.set_color(C_TEXT)
        ax.grid(True, alpha=0.25)
        plt.tight_layout()
        st.pyplot(fig)
        plt.close()
    else:
        st.info("Historial insuficiente para mostrar tendencia gráfica.")

    section_header("Todos tus productos")
    busqueda = st.text_input("Buscar producto por nombre", placeholder="Escribe para filtrar...",
                             label_visibility="collapsed")
    productos_filtrados = [p for p in productos if busqueda.lower() in p.lower()] if busqueda else productos

    LIMITE_TABLA = 300
    tabla = []
    for p in productos_filtrados[:LIMITE_TABLA]:
        s = st.session_state['stock_data'].get(p)
        if s is None:
            s = int(df_agg[df_agg['Producto']==p]['Cantidad'].mean()*2)
        i = calcular_estado_inventario(payload, p, s)
        if i:
            tabla.append({
                'Producto': p[:50],
                'Stock actual': s,
                'Demanda estimada (1 sem.)': i['demanda_predicha'],
                'Estado': {'critico':'Crítico','atencion':'Atención',
                          'exceso':'Exceso','optimo':'Óptimo'}[i['estado']],
            })

    if len(productos_filtrados) > LIMITE_TABLA:
        st.caption(f"Mostrando los primeros {LIMITE_TABLA} de {len(productos_filtrados):,} productos encontrados. "
                  "Usa el buscador para encontrar uno específico.")
    elif busqueda:
        st.caption(f"{len(productos_filtrados):,} producto(s) encontrado(s) de {len(productos):,} en total.")

    st.dataframe(pd.DataFrame(tabla), use_container_width=True, height=380)
    footer_credits()


# ============================================================
#  VISTA: PREDICCIONES
# ============================================================
def view_predicciones():
    payload = st.session_state['payload']
    df_agg  = payload['df_agg']

    if 'stock_data' not in st.session_state:
        st.session_state['stock_data'] = {}

    col_txt, col_h = st.columns([3, 1])
    with col_txt:
        st.markdown("Planifica tus próximos pedidos antes de que falte stock. "
                   "Los productos se ordenan por urgencia: primero los que necesitas reponer más pronto.")
    with col_h:
        horizonte = st.selectbox("Horizonte de planificación", [1, 2, 3, 4],
                                 format_func=lambda x: f"Próximas {x} semana(s)")

    productos = df_agg['Producto'].unique()
    eventos = []
    for p in productos[:80]:
        s = st.session_state['stock_data'].get(p)
        if s is None:
            s = int(df_agg[df_agg['Producto']==p]['Cantidad'].mean()*2)
            st.session_state['stock_data'][p] = s
        info = calcular_estado_inventario(payload, p, s, semanas_adelante=horizonte)
        if info and info['unidades_sugeridas'] > 0:
            info['producto'] = p
            eventos.append(info)

    eventos = sorted(eventos, key=lambda x: x['fecha_sugerida_dt'])

    col1, col2, col3 = st.columns(3)
    with col1:
        metric_card("Pedidos pendientes", f"{len(eventos)}", "productos a reponer", "warning")
    with col2:
        proximos_7d = [e for e in eventos if (e['fecha_sugerida_dt'] - datetime.now()).days <= 7]
        metric_card("Urgentes (7 días)", f"{len(proximos_7d)}", "requieren pedido esta semana", "danger")
    with col3:
        total_unidades = sum(e['unidades_sugeridas'] for e in eventos)
        metric_card("Unidades a pedir", f"{total_unidades:,}", "en total, todos los productos", "info")

    section_header("Calendario de reposición")
    if not eventos:
        st.markdown('<div class="alert-success">No hay pedidos pendientes por ahora. Tu inventario está en buen estado.</div>',
                   unsafe_allow_html=True)
        footer_credits()
        return

    meses_es = {1:'ENE',2:'FEB',3:'MAR',4:'ABR',5:'MAY',6:'JUN',
               7:'JUL',8:'AGO',9:'SEP',10:'OCT',11:'NOV',12:'DIC'}

    for e in eventos[:25]:
        dias_restantes = (e['fecha_sugerida_dt'] - datetime.now()).days
        if dias_restantes <= 3:
            color_box, bg_box = C_DANGER, C_DANGER_BG
        elif dias_restantes <= 10:
            color_box, bg_box = C_WARNING, C_WARNING_BG
        else:
            color_box, bg_box = C_ACCENT_SOFT, C_BG_SOFT

        col_fecha, col_info, col_btn = st.columns([0.9, 4, 1])
        with col_fecha:
            st.markdown(f"""
            <div class="repo-date-box" style="background:{bg_box}; color:{color_box};">
                <div class="day">{e['fecha_sugerida_dt'].day}</div>
                <div class="mon">{meses_es[e['fecha_sugerida_dt'].month]}</div>
            </div>
            """, unsafe_allow_html=True)
        with col_info:
            urgencia = "Urgente" if dias_restantes <= 3 else ("Próximamente" if dias_restantes <= 10 else "Planificado")
            st.markdown(f"""
            <div style="padding-top:4px;">
                <b style="color:{C_TEXT};">{e['producto'][:55]}</b><br>
                <span style="font-size:13px; color:{C_TEXT_DIM};">
                    {urgencia} · Pedir <b style="color:{C_TEXT};">{e['unidades_sugeridas']} unidades</b> ·
                    Cobertura actual: {e['dias_cobertura']} días
                </span>
            </div>
            """, unsafe_allow_html=True)
        with col_btn:
            if st.button("Ver producto", key=f"repo_{e['producto']}", use_container_width=True):
                st.session_state['producto_detalle'] = e['producto']
                st.session_state['_force_inventario'] = True
                st.rerun()
        st.markdown("<div style='height:2px;'></div>", unsafe_allow_html=True)

    footer_credits()


# ============================================================
#  VISTA: REPORTES
# ============================================================
def view_reportes():
    payload = st.session_state['payload']
    df_agg  = payload['df_agg']

    if 'stock_data' not in st.session_state:
        st.session_state['stock_data'] = {}

    st.markdown("Descarga un resumen con el estado de tu inventario, las alertas activas "
               "y las predicciones de demanda, listo para compartir con tu equipo de compras o tus proveedores.")

    productos = df_agg['Producto'].unique()
    filas = []
    for p in productos[:200]:
        s = st.session_state['stock_data'].get(p)
        if s is None:
            s = int(df_agg[df_agg['Producto']==p]['Cantidad'].mean()*2)
        i = calcular_estado_inventario(payload, p, s)
        if i:
            filas.append({
                'Producto': p,
                'Stock actual': s,
                'Demanda estimada (próx. semana)': i['demanda_predicha'],
                'Stock de seguridad': i['stock_seguridad'],
                'Punto de reposición': i['punto_reposicion'],
                'Días de cobertura': i['dias_cobertura'],
                'Estado': {'critico':'Crítico','atencion':'Atención',
                          'exceso':'Exceso','optimo':'Óptimo'}[i['estado']],
                'Unidades sugeridas a pedir': i['unidades_sugeridas'],
            })

    df_reporte = pd.DataFrame(filas)
    section_header("Vista previa del reporte")
    st.dataframe(df_reporte, use_container_width=True, height=340)

    col1, col2, col3 = st.columns(3)
    with col1:
        excel_buffer = generar_reporte_excel(df_reporte)
        st.download_button("📊 Descargar en Excel (.xlsx)", data=excel_buffer,
                          file_name=f"reporte_inventario_{datetime.now().strftime('%Y%m%d')}.xlsx",
                          mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                          use_container_width=True, type="primary")
    with col2:
        csv = df_reporte.to_csv(index=False).encode('utf-8-sig')
        st.download_button("Descargar en CSV", data=csv,
                          file_name=f"reporte_inventario_{datetime.now().strftime('%Y%m%d')}.csv",
                          mime='text/csv', use_container_width=True)
    with col3:
        resumen_txt = generar_resumen_texto(df_reporte)
        st.download_button("Resumen ejecutivo (TXT)", data=resumen_txt.encode('utf-8'),
                          file_name=f"resumen_ejecutivo_{datetime.now().strftime('%Y%m%d')}.txt",
                          mime='text/plain', use_container_width=True)
    footer_credits()


# ============================================================
#  VISTA: MODELO (nueva pestaña)
# ============================================================
def view_modelo():
    payload = st.session_state['payload']
    metricas = payload['metricas']
    modelo    = payload['modelo']
    features  = payload['features']  # nombres originales en inglés

    st.markdown("## Rendimiento del modelo predictivo")
    st.caption("Métricas calculadas sobre el 20% de datos de prueba del dataset Online Retail 2010‑2011.")

    col1, col2, col3 = st.columns(3)
    with col1:
        metric_card("MAE", f"{metricas['mae']:.2f}", "Error absoluto medio (unidades)")
    with col2:
        metric_card("RMSE", f"{metricas['rmse']:.2f}", "Raíz del error cuadrático medio")
    with col3:
        metric_card("R²", f"{metricas['r2']:.2f}", "Coeficiente de determinación")

    section_header("Importancia de variables")
    importancias = pd.Series(modelo.feature_importances_, index=features).sort_values()
    fig, ax = plt.subplots(figsize=(10, 4))
    importancias.plot.barh(ax=ax, color=C_ACCENT_SOFT, edgecolor=C_BG)
    ax.set_title('¿Qué variables influyen más en la demanda?', fontweight='bold')
    ax.spines[['top', 'right']].set_visible(False)
    for i, v in enumerate(importancias):
        ax.text(v + 0.005, i, f'{v:.3f}', va='center', fontsize=9, color=C_TEXT)
    plt.tight_layout()
    st.pyplot(fig)
    plt.close()

    st.caption("Variables ordenadas por importancia relativa según Random Forest.")

    footer_credits()


# ============================================================
#  CONFIGURACIÓN (solo informativa en esta versión)
# ============================================================
def view_configuracion():
    st.markdown("## Configuración y datos")
    st.info("El modelo ya está pre‑entrenado con el dataset **Online Retail 2010‑2011**. "
            "No es necesario subir archivos. Si deseas actualizar los datos o el modelo, "
            "reemplaza los archivos .pkl y .csv en el repositorio y luego presiona 'Recargar modelo' en la barra lateral.")
    st.markdown("---")
    st.markdown("**Archivos utilizados:**")
    st.code("""
modelo_rf_optimizado.pkl
scaler.pkl
le_country.pkl
features.pkl
df_agg_online_retail.csv
    """)
    footer_credits()


# ============================================================
#  MAIN
# ============================================================

def main():
    # Estado inicial de sesión
    if 'logged_in' not in st.session_state:
        st.session_state['logged_in'] = False
    if not st.session_state['logged_in']:
        show_login()
        return

    # Carga automática del modelo al iniciar sesión (si aún no se ha cargado)
    if 'payload' not in st.session_state:
        with st.spinner('Cargando modelo entrenado y datos históricos...'):
            payload = cargar_modelo_y_datos()
            st.session_state['payload'] = payload
            st.session_state['stock_data'] = {}  # stock simulado
        st.success("Sistema listo. Navega por las pestañas para explorar tu inventario.")

    show_sidebar()
    show_topbar()

    # Si el usuario fuerza ir a inventario desde un botón de detalle
    if st.session_state.pop('_force_inventario', False):
        # Cambiamos el índice de pestaña activa manualmente (no soportado directamente,
        # pero podemos forzar la vista inventario en una segunda llamada).
        # Para simplificar, redirigimos a la pestaña Inventario recargando con parámetro.
        st.query_params['tab'] = 'inventario'
        st.rerun()

    # Definir pestañas (5 pestañas)
    labels = ["Inicio", "Mi Inventario", "Predicciones", "Reportes", "Modelo"]
    tabs = st.tabs(labels)

    # Determinar pestaña activa (por defecto Inicio)
    active_tab = st.query_params.get('tab', 'inicio')
    tab_index = 0
    if active_tab == 'inventario':
        tab_index = 1
    elif active_tab == 'predicciones':
        tab_index = 2
    elif active_tab == 'reportes':
        tab_index = 3
    elif active_tab == 'modelo':
        tab_index = 4

    with tabs[0]: view_inicio()
    with tabs[1]: view_inventario()
    with tabs[2]: view_predicciones()
    with tabs[3]: view_reportes()
    with tabs[4]: view_modelo()


if __name__ == "__main__":
    main()
